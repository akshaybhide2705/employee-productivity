from flask import Flask, render_template, request, redirect
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
excel_file = 'productivity_data.xlsx'

def save_to_excel(data):
    # If file doesn't exist, create it and write headers
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Employee Name", "Hours", "Activity", "Weightage", "Incident Number"])
    else:
        wb = load_workbook(excel_file)
        ws = wb.active

    ws.append(data)
    wb.save(excel_file)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date = request.form['date']
        name = request.form['name']
        hours = request.form['hours']
        activity = request.form['activity']
        weightage = request.form['weightage']
        incident = request.form['incident']

        save_to_excel([date, name, hours, activity, weightage, incident])
        return redirect('/')

    return render_template('form.html')

if __name__ == '__main__':
    app.run(debug=True)
